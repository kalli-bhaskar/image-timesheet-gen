from __future__ import annotations

from datetime import datetime
from calendar import monthrange
import json
import os
from pathlib import Path
import re
from shutil import copyfile
from typing import Any

from flask import Flask, jsonify, request, send_file
from openpyxl import load_workbook

try:
    import psycopg
except Exception:  # pragma: no cover - optional dependency in local-only runs
    psycopg = None

ROOT = Path(__file__).resolve().parent
UPLOADS_DIR = ROOT / 'uploads'
UPLOADS_DIR.mkdir(exist_ok=True)

TEMPLATE_CANDIDATES = [
    ROOT / 'Timesheet format March 1-15.xlsx',
    ROOT / 'Timesheet format.xlsx',
    ROOT / 'Timesheet_format.xlsx',
]

COUNTY_TAGS = {
    'fairfield': 'LCT',
    'franklin': 'CLB',
    'licking': 'NBY',
}

CITY_TO_COUNTY = {
    'columbus': 'franklin',
    'lancaster': 'fairfield',
    'newark': 'licking',
}

TAG_TO_DATA_CENTER = {
    'CLB': 'Columbus',
    'LCT': 'Lancaster',
    'NBY': 'Newark',
}

PROFILE_DEFAULTS = {
    'city_state': 'Columbus, OH',
    'customer': 'N/A',
    'full_name': '',
    'classification': 'N/A',
    'hourly_rate': 0,
    'per_diem': 'N/A',
    'accommodation_allowance': 'N/A',
    'stn_accommodation': 'N/A',
    'stn_rental': 'N/A',
    'stn_gas': 'N/A',
    'comment': '',
}

HEADER_ALIASES = {
    'date': 'Date',
    'data center location': 'Data Center Location',
    'city/state': 'City/State',
    'customer': 'Customer',
    'candidate name': 'Candidate Name',
    'classification employee/subcon': 'Classification Employee/Subcon',
    'hourly rate': 'Hourly rate',
    'per diem': 'Per Diem',
    'accommodation allowance': 'Accommodation Allowance',
    'time in': 'Time In',
    'time out': 'Time Out',
    'total hours': 'Total Hours',
    'hours in decimal': 'Hours in Decimal',
    'stn accommodation yes/no': 'STN Accommodation Yes/No',
    'stn rental yes/no': 'STN Rental Yes/ No',
    'stn gas yes/no': 'STN Gas Yes/ No',
    'comment': 'Comment',
}

app = Flask(__name__)


def load_local_env(env_path: Path) -> None:
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding='utf-8').splitlines():
        line = raw_line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        key, value = line.split('=', 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_local_env(ROOT / '.env.local')

DATABASE_URL = os.getenv('DATABASE_URL') or os.getenv('POSTGRES_URL') or os.getenv('POSTGRES_PRISMA_URL')
DB_ENABLED = bool(DATABASE_URL and psycopg)
DB_BOOTSTRAPPED = False
PAYROLL_SECOND_PERIOD_START_DAY = int(os.getenv('PAYROLL_SECOND_PERIOD_START_DAY', '16'))


def get_payroll_second_period_start_day(raw_day: Any = None) -> int:
    try:
        candidate = raw_day if raw_day not in (None, '') else PAYROLL_SECOND_PERIOD_START_DAY
        day = int(candidate or 16)
    except (TypeError, ValueError):
        day = 16
    if day < 2 or day > 28:
        return 16
    return day


def get_payroll_period(shift_dt: datetime, second_period_start_day: Any = None) -> str:
    second_start = get_payroll_second_period_start_day(second_period_start_day)
    first_end = second_start - 1
    year = shift_dt.year
    month = shift_dt.month
    day = shift_dt.day

    if day <= first_end:
        start = datetime(year, month, 1).date().isoformat()
        end = datetime(year, month, first_end).date().isoformat()
        return f'{start} to {end}'

    month_last_day = monthrange(year, month)[1]
    start = datetime(year, month, second_start).date().isoformat()
    end = datetime(year, month, month_last_day).date().isoformat()
    return f'{start} to {end}'


def get_db_connection() -> Any:
    if not DB_ENABLED:
        return None
    return psycopg.connect(DATABASE_URL)


def initialize_database() -> None:
    global DB_BOOTSTRAPPED
    if DB_BOOTSTRAPPED:
        return
    if not DB_ENABLED:
        DB_BOOTSTRAPPED = True
        return
    schema_path = ROOT / 'database' / 'schema.sql'
    if not schema_path.exists():
        return
    schema_sql = schema_path.read_text(encoding='utf-8')
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(schema_sql)
        conn.commit()
    DB_BOOTSTRAPPED = True


def normalize_header(value: Any) -> str:
    if value is None:
        return ''
    cleaned = str(value).replace('\xa0', ' ')
    cleaned = cleaned.replace(' /', '/').replace('/ ', '/')
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.strip().lower()


def get_template_path(shift_dt: datetime | None = None) -> Path:
    if shift_dt:
        march_template = ROOT / 'Timesheet format March 1-15.xlsx'
        if march_template.exists() and shift_dt.date().month == 3 and 1 <= shift_dt.date().day <= 15:
            return march_template
    for candidate in TEMPLATE_CANDIDATES:
        if candidate.exists():
            return candidate
    raise FileNotFoundError('No timesheet template was found in the workspace root.')


def normalize_location_text(value: str | None) -> str:
    text = re.sub(r'[^a-z\s]', ' ', str(value or '').lower())
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def canonicalize_county(value: str | None) -> str:
    normalized = normalize_location_text(value)
    if normalized.endswith(' county'):
        normalized = normalized[: -len(' county')].strip()
    if normalized in COUNTY_TAGS:
        return normalized
    for county_key in COUNTY_TAGS:
        if county_key in normalized:
            return county_key
    for city_name, county_key in CITY_TO_COUNTY.items():
        if city_name in normalized:
            return county_key
    return normalized


def determine_location_tag(county: str | None) -> str:
    canonical_county = canonicalize_county(county)
    if not canonical_county:
        return ''
    return COUNTY_TAGS.get(canonical_county, '')


def sanitize_tag(value: str | None, valid_tags: set[str] | None = None) -> str | None:
    """Return value if it looks like a valid location tag code, else None."""
    if not value:
        return None
    upper = str(value).upper().strip()
    # 'AUTO' and other placeholder strings are not real tags
    if not re.match(r'^[A-Z]{2,8}$', upper):
        return None
    if valid_tags is not None:
        return upper if upper in valid_tags else None
    return upper


def get_header_map(worksheet: Any) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for cell in worksheet[1]:
        normalized = normalize_header(cell.value)
        if not normalized:
            continue
        header_map[normalized] = cell.column
    return header_map


def get_row_for_date(worksheet: Any, shift_dt: datetime) -> int:
    for row in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row, column=1).value
        if isinstance(value, datetime) and value.date() == shift_dt.date():
            return row
        if hasattr(value, 'date') and value.date() == shift_dt.date():
            return row
    raise ValueError(f'No row found for {shift_dt.date().isoformat()} in the selected workbook.')


def calculate_hours(clock_in: datetime | None, clock_out: datetime | None) -> tuple[str | None, float | None]:
    if not clock_in or not clock_out:
        return None, None
    delta = clock_out - clock_in
    total_seconds = max(delta.total_seconds(), 0)
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f'{hours:02d}:{minutes:02d}', round(total_seconds / 3600, 2)


def extract_profile(form_data: Any) -> dict[str, Any]:
    profile = dict(PROFILE_DEFAULTS)
    for key in profile:
        incoming = form_data.get(key)
        if incoming is not None and incoming != '':
            profile[key] = incoming
    try:
        profile['hourly_rate'] = float(profile['hourly_rate'])
    except (TypeError, ValueError):
        profile['hourly_rate'] = PROFILE_DEFAULTS['hourly_rate']
    return profile


def write_shift_to_workbook(
    clock_in_result: dict[str, Any] | None,
    clock_out_result: dict[str, Any] | None,
    profile: dict[str, Any],
) -> Path:
    shift_dt = None
    if clock_in_result and clock_in_result['timestamp']:
        shift_dt = datetime.fromisoformat(clock_in_result['timestamp'])
    elif clock_out_result and clock_out_result['timestamp']:
        shift_dt = datetime.fromisoformat(clock_out_result['timestamp'])
    if not shift_dt:
        raise ValueError('Unable to determine a shift date from OCR output.')

    output_path = UPLOADS_DIR / 'timesheet_latest.xlsx'
    if not output_path.exists():
        template_path = get_template_path(shift_dt)
        copyfile(template_path, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook.active
    row = get_row_for_date(sheet, shift_dt)
    header_map = get_header_map(sheet)

    def column_for(name: str) -> int:
        column = header_map.get(normalize_header(name))
        if not column:
            raise KeyError(f'Missing expected worksheet column: {name}')
        return column

    time_in_dt = datetime.fromisoformat(clock_in_result['timestamp']) if clock_in_result and clock_in_result['timestamp'] else None
    time_out_dt = datetime.fromisoformat(clock_out_result['timestamp']) if clock_out_result and clock_out_result['timestamp'] else None
    hours_text, hours_decimal = calculate_hours(time_in_dt, time_out_dt)
    location_tag = ''
    for result in (clock_in_result, clock_out_result):
        if not result:
            continue
        if result.get('location_tag'):
            location_tag = result['location_tag']
            break
        if result.get('county'):
            mapped = determine_location_tag(str(result['county']))
            if mapped:
                location_tag = mapped
            break

    values = {
        'Data Center Location': location_tag,
        'City/State': profile['city_state'],
        'Customer': profile['customer'],
        'Candidate Name': profile['full_name'],
        'Classification Employee/Subcon': profile['classification'],
        'Hourly rate': profile['hourly_rate'],
        'Per Diem': profile['per_diem'],
        'Accommodation Allowance': profile['accommodation_allowance'],
        'Time In': time_in_dt.time() if time_in_dt else None,
        'Time Out': time_out_dt.time() if time_out_dt else None,
        'Total Hours': hours_text,
        'Hours in Decimal': hours_decimal,
        'STN Accommodation Yes/No': profile['stn_accommodation'],
        'STN Rental Yes/ No': profile['stn_rental'],
        'STN Gas Yes/ No': profile['stn_gas'],
        'Comment': profile['comment'],
    }

    for header_name, value in values.items():
        sheet.cell(row=row, column=column_for(header_name), value=value)

    workbook.save(output_path)
    return output_path


def normalize_meta_result(meta: Any, label: str) -> dict[str, Any] | None:
    if not meta:
        return None
    if not isinstance(meta, dict):
        raise ValueError(f'{label} must be an object when provided.')

    timestamp = meta.get('timestamp')
    if timestamp:
        try:
            datetime.fromisoformat(str(timestamp))
        except ValueError as exc:
            raise ValueError(f'{label}.timestamp must be ISO format, e.g. 2026-03-05T09:32:59.') from exc

    county_value = meta.get('county')
    canonical_county = canonicalize_county(str(county_value)) if county_value else ''
    county = canonical_county.title() if canonical_county else county_value
    location_tag = str(meta.get('location_tag') or '').strip().upper()
    if not location_tag and county:
        location_tag = determine_location_tag(str(county))
    data_center_location = str(meta.get('data_center_location') or '').strip()
    if not data_center_location and location_tag:
        data_center_location = TAG_TO_DATA_CENTER.get(location_tag, '')

    return {
        'filename': meta.get('filename') or label,
        'ocr_text': meta.get('ocr_text') or '',
        'timestamp': timestamp,
        'date': meta.get('date'),
        'time': meta.get('time'),
        'county': county,
        'location_tag': location_tag or '',
        'data_center_location': data_center_location,
        'crop': meta.get('crop') or {'strategy': 'metadata-only'},
        'file_url': str(meta.get('file_url') or '').strip() or None,
    }


def extract_submission_actor(source: Any) -> dict[str, str]:
    def get_value(key: str) -> str:
        value = source.get(key) if hasattr(source, 'get') else None
        if value is None:
            return ''
        return str(value).strip()

    return {
        'employee_email': get_value('employee_email').lower(),
        'employee_name': get_value('employee_name'),
        'manager_email': get_value('manager_email').lower(),
        'customer': get_value('customer') or get_value('company_name'),
        'work_location_tag': get_value('work_location_tag').upper(),
    }


def persist_shift_to_database(
    clock_in_result: dict[str, Any] | None,
    clock_out_result: dict[str, Any] | None,
    profile: dict[str, Any],
    actor: dict[str, str],
) -> dict[str, Any]:
    if not DB_ENABLED:
        return {'enabled': False, 'saved': False, 'reason': 'database_not_configured'}

    employee_email = actor.get('employee_email', '').strip().lower()
    if not employee_email:
        return {'enabled': True, 'saved': False, 'reason': 'employee_email_missing'}

    time_in_dt = datetime.fromisoformat(clock_in_result['timestamp']) if clock_in_result and clock_in_result.get('timestamp') else None
    time_out_dt = datetime.fromisoformat(clock_out_result['timestamp']) if clock_out_result and clock_out_result.get('timestamp') else None

    shift_dt = None
    for dt in (time_in_dt, time_out_dt):
        if dt:
            shift_dt = dt
            break
    if not shift_dt:
        return {'enabled': True, 'saved': False, 'reason': 'missing_shift_datetime'}

    hours_text, hours_decimal = calculate_hours(time_in_dt, time_out_dt)

    location_tag = ''
    for result in (clock_in_result, clock_out_result):
        if result and result.get('location_tag'):
            location_tag = str(result['location_tag']).upper()
            break

    manager_email = actor.get('manager_email', '').lower()
    customer_name = actor.get('customer') or profile.get('customer') or None
    employee_name = actor.get('employee_name') or profile.get('full_name') or employee_email.split('@')[0]
    work_location_tag = actor.get('work_location_tag') or None
    status = 'completed' if time_in_dt and time_out_dt else 'clocked_in'
    payroll_start_day_raw = profile.get('payroll_second_period_start_day')
    try:
        payroll_start_day = int(payroll_start_day_raw) if payroll_start_day_raw is not None else 16
    except (TypeError, ValueError):
        payroll_start_day = 16
    payroll_start_day = max(2, min(28, payroll_start_day))
    clock_in_photo_url = str(clock_in_result.get('file_url') or '').strip() or None if clock_in_result else None
    clock_out_photo_url = str(clock_out_result.get('file_url') or '').strip() or None if clock_out_result else None

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                # Fetch valid tags once to avoid FK violations from placeholder values like 'AUTO'
                cur.execute('SELECT tag FROM location_tags')
                valid_tags: set[str] = {row[0] for row in cur.fetchall()}
                work_location_tag = sanitize_tag(work_location_tag, valid_tags)
                location_tag = sanitize_tag(location_tag, valid_tags) or ''

                manager_id = None
                if manager_email:
                    cur.execute(
                        """
                        insert into app_users (
                          email, full_name, display_name, user_role, setup_complete, company_name, customer
                        )
                        values (%s, %s, %s, 'manager', true, %s, %s)
                        on conflict (email) do update
                        set company_name = coalesce(excluded.company_name, app_users.company_name),
                            customer = coalesce(excluded.customer, app_users.customer),
                            updated_at = now()
                        returning id
                        """,
                        (manager_email, manager_email.split('@')[0], manager_email.split('@')[0], customer_name, customer_name),
                    )
                    row = cur.fetchone()
                    manager_id = row[0] if row else None

                cur.execute(
                    """
                    insert into app_users (
                      email, full_name, display_name, user_role, setup_complete,
                      manager_id, company_name, city_state, customer, classification,
                      hourly_rate, per_diem, accommodation_allowance,
                                            stn_accommodation, stn_rental, stn_gas, work_location_tag, payroll_second_period_start_day
                    )
                                        values (%s, %s, %s, 'employee', true, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    on conflict (email) do update
                    set
                      full_name = excluded.full_name,
                      display_name = excluded.display_name,
                      manager_id = coalesce(excluded.manager_id, app_users.manager_id),
                      company_name = coalesce(excluded.company_name, app_users.company_name),
                      city_state = excluded.city_state,
                      customer = excluded.customer,
                      classification = excluded.classification,
                      hourly_rate = excluded.hourly_rate,
                      per_diem = excluded.per_diem,
                      accommodation_allowance = excluded.accommodation_allowance,
                      stn_accommodation = excluded.stn_accommodation,
                      stn_rental = excluded.stn_rental,
                      stn_gas = excluded.stn_gas,
                                            payroll_second_period_start_day = coalesce(excluded.payroll_second_period_start_day, app_users.payroll_second_period_start_day),
                      work_location_tag = coalesce(excluded.work_location_tag, app_users.work_location_tag),
                      updated_at = now()
                                        returning id, manager_id, payroll_second_period_start_day
                    """,
                    (
                        employee_email,
                        employee_name,
                        employee_name,
                        manager_id,
                        customer_name,
                        profile.get('city_state'),
                        customer_name,
                        profile.get('classification'),
                        float(profile.get('hourly_rate') or 0),
                        profile.get('per_diem'),
                        profile.get('accommodation_allowance'),
                        profile.get('stn_accommodation'),
                        profile.get('stn_rental'),
                        profile.get('stn_gas'),
                        work_location_tag,
                        payroll_start_day,
                    ),
                )
                employee_row = cur.fetchone()
                employee_id = employee_row[0]
                resolved_manager_id = employee_row[1] or manager_id
                user_period_start_day = get_payroll_second_period_start_day(employee_row[2] if len(employee_row) > 2 else None)
                payroll_period = get_payroll_period(shift_dt, user_period_start_day)

                entry_row = None
                if time_out_dt:
                    # Complete the latest open clock-in row first to avoid duplicates.
                    cur.execute(
                        """
                        update timesheet_entries
                        set
                          manager_id = %s,
                          shift_date = %s,
                          payroll_period = %s,
                          status = %s,
                          data_center_location = %s,
                          location_source = %s,
                          time_in = %s,
                          time_out = %s,
                          total_hours_text = %s,
                          hours_decimal = %s,
                          clock_in_photo_url = coalesce(%s, clock_in_photo_url),
                          clock_out_photo_url = coalesce(%s, clock_out_photo_url),
                          clock_in_meta = %s::jsonb,
                          clock_out_meta = %s::jsonb,
                          comment = %s,
                          updated_at = now()
                        where id = (
                          select id from timesheet_entries
                          where employee_id = %s and status = 'clocked_in' and time_out is null
                          order by created_at desc
                          limit 1
                        )
                        returning id
                        """,
                        (
                            resolved_manager_id,
                            shift_dt.date(),
                            payroll_period,
                            'completed',
                            location_tag or None,
                            'backend_ocr',
                            time_in_dt,
                            time_out_dt,
                            hours_text,
                            float(hours_decimal or 0),
                            clock_in_photo_url,
                            clock_out_photo_url,
                            json.dumps(clock_in_result) if clock_in_result else None,
                            json.dumps(clock_out_result) if clock_out_result else None,
                            profile.get('comment') or '',
                            employee_id,
                        ),
                    )
                    entry_row = cur.fetchone()

                if not entry_row:
                    cur.execute(
                        """
                        insert into timesheet_entries (
                          employee_id, manager_id, shift_date, payroll_period, status,
                          data_center_location, location_source,
                          time_in, time_out, total_hours_text, hours_decimal,
                          clock_in_photo_url, clock_out_photo_url, clock_in_meta, clock_out_meta, comment
                        )
                        values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, %s)
                        returning id
                        """,
                        (
                            employee_id,
                            resolved_manager_id,
                            shift_dt.date(),
                            payroll_period,
                            status,
                            location_tag or None,
                            'backend_ocr',
                            time_in_dt,
                            time_out_dt,
                            hours_text,
                            float(hours_decimal or 0),
                            clock_in_photo_url,
                            clock_out_photo_url,
                            json.dumps(clock_in_result) if clock_in_result else None,
                            json.dumps(clock_out_result) if clock_out_result else None,
                            profile.get('comment') or '',
                        ),
                    )
                    entry_row = cur.fetchone()
            conn.commit()

        return {'enabled': True, 'saved': True, 'entry_id': str(entry_row[0]) if entry_row else ''}
    except Exception as exc:  # pragma: no cover - runtime env specific
        return {'enabled': True, 'saved': False, 'error': str(exc)}


@app.after_request
def add_cors_headers(response: Any) -> Any:
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET,POST,OPTIONS'
    return response


@app.before_request
def ensure_database_ready() -> None:
    if request.method == 'OPTIONS':
        return
    initialize_database()


@app.route('/health', methods=['GET'])
def health() -> Any:
    return jsonify({'ok': True, 'database_enabled': DB_ENABLED})


@app.route('/admin/clear-entries', methods=['POST'])
def clear_entries() -> Any:
    """Delete timesheet entries for one user (or a whole manager's team) from the DB.
    Body: { "email": "...", "scope": "self" | "team" }
    - scope=self  → delete entries for that email only (default)
    - scope=team  → delete entries for all employees managed by that email
    """
    payload = request.get_json(silent=True) or {}
    email = str(payload.get('email') or '').strip().lower()
    scope = str(payload.get('scope') or 'self').strip().lower()
    if not email:
        return jsonify({'error': 'email required'}), 400

    if not DB_ENABLED:
        return jsonify({'cleared': 0, 'reason': 'db_disabled'}), 200

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                if scope == 'team':
                    cur.execute(
                        """
                        delete from timesheet_entries
                        where employee_id in (
                          select id from app_users
                          where manager_id = (select id from app_users where email = %s)
                        )
                        """,
                        (email,),
                    )
                else:
                    cur.execute(
                        """
                        delete from timesheet_entries
                        where employee_id = (select id from app_users where email = %s)
                        """,
                        (email,),
                    )
                cleared = cur.rowcount
            conn.commit()
        return jsonify({'cleared': cleared}), 200
    except Exception as exc:
        return jsonify({'cleared': 0, 'error': str(exc)}), 200


@app.route('/user', methods=['POST'])
def upsert_user() -> Any:
    """Create or update a user record from the frontend profile/setup data."""
    payload = request.get_json(silent=True) or {}
    email = str(payload.get('email') or '').strip().lower()
    if not email:
        return jsonify({'error': 'email required'}), 400

    if not DB_ENABLED:
        return jsonify({'saved': False, 'reason': 'db_disabled'}), 200

    full_name = str(payload.get('full_name') or '').strip() or email.split('@')[0]
    display_name = str(payload.get('display_name') or full_name).strip()
    user_role = 'manager' if str(payload.get('user_role') or '').strip().lower() == 'manager' else 'employee'
    setup_complete = bool(payload.get('setup_complete', False))
    company_name = payload.get('company_name') or None
    city_state = payload.get('city_state') or None
    customer = payload.get('customer') or payload.get('company_name') or None
    classification = payload.get('classification') or None
    per_diem = payload.get('per_diem') or None
    accommodation_allowance = payload.get('accommodation_allowance') or None
    stn_accommodation = payload.get('stn_accommodation') or None
    stn_rental = payload.get('stn_rental') or None
    stn_gas = payload.get('stn_gas') or None
    manager_email = str(payload.get('manager_email') or '').strip().lower() or None
    try:
        hourly_rate = float(payload.get('hourly_rate') or 0)
    except (TypeError, ValueError):
        hourly_rate = 0.0
    raw_tag = str(payload.get('work_location_tag') or '').strip().upper() or None
    location_enforcement_enabled = bool(payload.get('location_enforcement_enabled') or False)
    payroll_second_period_start_day = get_payroll_second_period_start_day(payload.get('payroll_second_period_start_day'))

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT tag FROM location_tags')
                valid_tags = {row[0] for row in cur.fetchall()}
                work_location_tag = raw_tag if raw_tag in valid_tags else None

                manager_id = None
                if manager_email:
                    cur.execute('SELECT id FROM app_users WHERE email = %s', (manager_email,))
                    mrow = cur.fetchone()
                    manager_id = mrow[0] if mrow else None

                cur.execute(
                    """
                    INSERT INTO app_users (
                      email, full_name, display_name, user_role, setup_complete,
                      company_name, city_state, customer, classification,
                      hourly_rate, per_diem, accommodation_allowance,
                      stn_accommodation, stn_rental, stn_gas,
                                            work_location_tag, manager_id, location_enforcement_enabled, payroll_second_period_start_day
                    )
                                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (email) DO UPDATE SET
                      full_name             = EXCLUDED.full_name,
                      display_name          = EXCLUDED.display_name,
                      user_role             = EXCLUDED.user_role,
                      setup_complete        = EXCLUDED.setup_complete,
                      company_name          = COALESCE(EXCLUDED.company_name, app_users.company_name),
                      city_state            = COALESCE(EXCLUDED.city_state, app_users.city_state),
                      customer              = COALESCE(EXCLUDED.customer, app_users.customer),
                      classification        = COALESCE(EXCLUDED.classification, app_users.classification),
                      hourly_rate           = EXCLUDED.hourly_rate,
                      per_diem              = COALESCE(EXCLUDED.per_diem, app_users.per_diem),
                      accommodation_allowance = COALESCE(EXCLUDED.accommodation_allowance, app_users.accommodation_allowance),
                      stn_accommodation     = COALESCE(EXCLUDED.stn_accommodation, app_users.stn_accommodation),
                      stn_rental            = COALESCE(EXCLUDED.stn_rental, app_users.stn_rental),
                      stn_gas               = COALESCE(EXCLUDED.stn_gas, app_users.stn_gas),
                      work_location_tag     = COALESCE(EXCLUDED.work_location_tag, app_users.work_location_tag),
                      manager_id            = COALESCE(EXCLUDED.manager_id, app_users.manager_id),
                      location_enforcement_enabled = EXCLUDED.location_enforcement_enabled,
                                            payroll_second_period_start_day = EXCLUDED.payroll_second_period_start_day,
                      updated_at            = now()
                    RETURNING id
                    """,
                    (
                        email, full_name, display_name, user_role, setup_complete,
                        company_name, city_state, customer, classification,
                        hourly_rate, per_diem, accommodation_allowance,
                        stn_accommodation, stn_rental, stn_gas,
                        work_location_tag, manager_id, location_enforcement_enabled, payroll_second_period_start_day,
                    ),
                )
                urow = cur.fetchone()
            conn.commit()
        return jsonify({'saved': True, 'id': str(urow[0]) if urow else None}), 200
    except Exception as exc:
        return jsonify({'saved': False, 'error': str(exc)}), 200


@app.route('/user', methods=['GET'])
def get_user() -> Any:
    email = request.args.get('email', '').strip().lower()
    if not email:
        return jsonify({'error': 'email query param required'}), 400

    if not DB_ENABLED:
        return jsonify({'found': False, 'reason': 'db_disabled'}), 200

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT u.id, u.email, u.full_name, u.display_name, u.user_role, u.setup_complete,
                           u.company_name, u.city_state, u.work_location_tag, u.manager_id,
                           u.customer, u.classification, u.hourly_rate, u.per_diem,
                           u.accommodation_allowance, u.stn_accommodation, u.stn_rental, u.stn_gas,
                              m.email AS manager_email, u.location_enforcement_enabled, u.payroll_second_period_start_day
                    FROM app_users u
                    LEFT JOIN app_users m ON m.id = u.manager_id
                    WHERE u.email = %s
                    """,
                    (email,),
                )
                row = cur.fetchone()

        if not row:
            return jsonify({'found': False}), 200

        col_names = [
            'id', 'email', 'full_name', 'display_name', 'user_role', 'setup_complete',
            'company_name', 'city_state', 'work_location_tag', 'manager_id',
            'customer', 'classification', 'hourly_rate', 'per_diem',
            'accommodation_allowance', 'stn_accommodation', 'stn_rental', 'stn_gas',
            'manager_email', 'location_enforcement_enabled', 'payroll_second_period_start_day',
        ]
        user = dict(zip(col_names, row))
        user['id'] = str(user['id']) if user['id'] else None
        user['manager_id'] = str(user['manager_id']) if user['manager_id'] else None
        user['hourly_rate'] = float(user['hourly_rate'] or 0)
        return jsonify({'found': True, 'user': user}), 200
    except Exception as exc:
        return jsonify({'found': False, 'error': str(exc)}), 200


@app.route('/users', methods=['GET'])
def list_users() -> Any:
    """List users from DB with optional filters.
    Supported query params: email, manager_email, user_role, customer
    """
    if not DB_ENABLED:
        return jsonify({'users': [], 'reason': 'db_disabled'}), 200

    email = request.args.get('email', '').strip().lower()
    manager_email = request.args.get('manager_email', '').strip().lower()
    user_role = request.args.get('user_role', '').strip().lower()
    customer = request.args.get('customer', '').strip().lower()

    where_parts: list[str] = []
    params: list[Any] = []

    if email:
        where_parts.append('lower(u.email) = %s')
        params.append(email)
    if manager_email:
        where_parts.append("lower(coalesce(m.email, '')) = %s")
        params.append(manager_email)
    if user_role:
        where_parts.append('lower(u.user_role) = %s')
        params.append(user_role)
    if customer:
        where_parts.append("lower(coalesce(nullif(u.customer, ''), u.company_name, '')) = %s")
        params.append(customer)

    where_sql = f" where {' and '.join(where_parts)}" if where_parts else ''

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      u.id,
                      u.email,
                      u.full_name,
                      u.display_name,
                      u.user_role,
                      u.setup_complete,
                      u.company_name,
                      coalesce(nullif(u.customer, ''), u.company_name, '') as customer,
                      u.city_state,
                      u.work_location_tag,
                      m.email as manager_email,
                      u.classification,
                      u.hourly_rate,
                      u.per_diem,
                      u.accommodation_allowance,
                      u.stn_accommodation,
                      u.stn_rental,
                      u.stn_gas,
                                            u.location_enforcement_enabled,
                                            u.payroll_second_period_start_day
                    FROM app_users u
                    LEFT JOIN app_users m ON m.id = u.manager_id
                    {where_sql}
                    ORDER BY u.full_name ASC, u.email ASC
                    """,
                    tuple(params),
                )
                rows = cur.fetchall()

        users = []
        for row in rows:
            users.append(
                {
                    'id': str(row[0]) if row[0] else None,
                    'email': row[1],
                    'full_name': row[2],
                    'display_name': row[3],
                    'user_role': row[4],
                    'setup_complete': bool(row[5]),
                    'company_name': row[6] or '',
                    'customer': row[7] or '',
                    'city_state': row[8] or '',
                    'work_location_tag': row[9] or '',
                    'manager_email': row[10] or '',
                    'classification': row[11] or '',
                    'hourly_rate': float(row[12] or 0),
                    'per_diem': row[13] or '',
                    'accommodation_allowance': row[14] or '',
                    'stn_accommodation': row[15] or '',
                    'stn_rental': row[16] or '',
                    'stn_gas': row[17] or '',
                    'location_enforcement_enabled': bool(row[18]) if row[18] is not None else False,
                    'payroll_second_period_start_day': int(row[19] or 16),
                }
            )

        return jsonify({'users': users}), 200
    except Exception as exc:
        return jsonify({'users': [], 'error': str(exc)}), 200


@app.route('/timesheet_entries', methods=['GET'])
def list_timesheet_entries() -> Any:
    if not DB_ENABLED:
        return jsonify({'entries': [], 'reason': 'db_disabled'}), 200

    employee_email = request.args.get('employee_email', '').strip().lower()
    manager_email = request.args.get('manager_email', '').strip().lower()
    status = request.args.get('status', '').strip().lower()
    limit_raw = request.args.get('limit', '').strip()

    where_parts: list[str] = []
    params: list[Any] = []

    if employee_email:
        where_parts.append('lower(e.email) = %s')
        params.append(employee_email)
    if manager_email:
        where_parts.append("lower(coalesce(m.email, '')) = %s")
        params.append(manager_email)
    if status:
        where_parts.append('lower(t.status) = %s')
        params.append(status)

    where_sql = f" where {' and '.join(where_parts)}" if where_parts else ''
    limit_sql = ''
    if limit_raw.isdigit():
        limit_sql = ' limit %s'
        params.append(int(limit_raw))

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    select
                      t.id,
                      e.email as employee_email,
                      e.full_name as employee_name,
                      t.shift_date,
                      t.data_center_location,
                      t.location_source,
                      e.city_state,
                      coalesce(nullif(e.customer, ''), e.company_name, '') as customer,
                      e.classification,
                      e.hourly_rate,
                      e.per_diem,
                      e.accommodation_allowance,
                      e.stn_accommodation,
                      e.stn_rental,
                      e.stn_gas,
                      t.time_in,
                      t.time_out,
                      t.total_hours_text,
                      t.hours_decimal,
                      t.comment,
                      t.clock_in_photo_url,
                      t.clock_out_photo_url,
                      t.clock_in_meta,
                      t.clock_out_meta,
                      t.payroll_period,
                      t.status,
                      t.created_at,
                      t.updated_at
                    from timesheet_entries t
                    join app_users e on e.id = t.employee_id
                    left join app_users m on m.id = t.manager_id
                    {where_sql}
                    order by t.created_at desc
                    {limit_sql}
                    """,
                    tuple(params),
                )
                rows = cur.fetchall()

        entries: list[dict[str, Any]] = []
        for row in rows:
            entries.append(
                {
                    'id': str(row[0]) if row[0] else None,
                    'employee_email': row[1] or '',
                    'employee_name': row[2] or '',
                    'date': row[3].isoformat() if row[3] else '',
                    'data_center_location': row[4] or '',
                    'location_source': row[5] or '',
                    'city_state': row[6] or '',
                    'customer': row[7] or '',
                    'classification': row[8] or '',
                    'hourly_rate': float(row[9] or 0),
                    'per_diem': row[10] or '',
                    'accommodation_allowance': row[11] or '',
                    'stn_accommodation': row[12] or '',
                    'stn_rental': row[13] or '',
                    'stn_gas': row[14] or '',
                    'time_in': row[15].isoformat() if row[15] else '',
                    'time_out': row[16].isoformat() if row[16] else '',
                    'total_hours': row[17] or '0:00',
                    'hours_decimal': float(row[18] or 0),
                    'comment': row[19] or '',
                    'clock_in_photo_url': row[20] or '',
                    'clock_out_photo_url': row[21] or '',
                    'clock_in_meta': row[22] or None,
                    'clock_out_meta': row[23] or None,
                    'payroll_period': row[24] or '',
                    'status': row[25] or '',
                    'created_date': row[26].isoformat() if row[26] else '',
                    'updated_date': row[27].isoformat() if row[27] else '',
                }
            )

        return jsonify({'entries': entries}), 200
    except Exception as exc:
        return jsonify({'entries': [], 'error': str(exc)}), 200


@app.route('/submit_shift_meta', methods=['POST'])
def submit_shift_meta() -> Any:
    payload = request.get_json(silent=True) or {}
    if not payload:
        return jsonify({'error': 'Send JSON body with clock_in, clock_out, and optional profile.'}), 400

    try:
        clock_in_result = normalize_meta_result(payload.get('clock_in'), 'clock_in')
        clock_out_result = normalize_meta_result(payload.get('clock_out'), 'clock_out')
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    if not clock_in_result and not clock_out_result:
        return jsonify({'error': 'Provide at least one of clock_in or clock_out metadata objects.'}), 400

    profile_input = payload.get('profile') or {}
    if not isinstance(profile_input, dict):
        return jsonify({'error': 'profile must be an object.'}), 400

    profile = extract_profile(profile_input)
    actor = extract_submission_actor(payload)

    # Always persist to DB first — independent of workbook generation.
    db_result = persist_shift_to_database(clock_in_result, clock_out_result, profile, actor)

    workbook_error = None
    workbook_path = None
    try:
        workbook_path = write_shift_to_workbook(clock_in_result, clock_out_result, profile)
    except (ValueError, KeyError, FileNotFoundError) as exc:
        workbook_error = str(exc)

    response = {
        'ok': True,
        'clock_in': clock_in_result,
        'clock_out': clock_out_result,
        'workbook_path': str(workbook_path) if workbook_path else None,
        'download_url': '/download/latest-timesheet' if workbook_path else None,
        'db': db_result,
    }
    if workbook_error:
        response['workbook_warning'] = workbook_error
    return jsonify(response)


@app.route('/download/latest-timesheet', methods=['GET'])
def download_latest_timesheet() -> Any:
    output_path = UPLOADS_DIR / 'timesheet_latest.xlsx'
    if not output_path.exists():
        return jsonify({'error': 'No generated workbook is available yet.'}), 404
    return send_file(output_path, as_attachment=True, download_name=output_path.name)


if __name__ == '__main__':
    initialize_database()
    app.run(host='0.0.0.0', port=8765, debug=True)